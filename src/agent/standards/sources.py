"""Raw standard readers (Phase 1): read the ORIGINAL standard workbooks into
plain raw-entry dicts — MERGED-RANGE-AWARE.

The standard workbooks use vertical cell merges for GROUP/hierarchy columns:
finance 金融行业数据安全分类分级标准指南 (B..F hierarchy + J remark, K dept),
shougang 关基-数据分类分级目录 (B..G hierarchy). A merged cell is ONE source
value owned by a RANGE of rows; ``MergedCellResolver`` expands it to every
covered cell while preserving the anchor and scope, so a leaf inherits its
group's definitions/remark without misattributing them as leaf-private.

Leaf columns (四级子类 / 内容 / 安全级别 / 分级 / 数据资源说明 / 数据来源) are
per-row cells (no merges). The three SAMPLE-source workbooks (部分金融数据 /
带分级分类… / 关基设施…用于测试) have no business-level merges and are NOT read
by this module.

Excel layout (verified against data/raw at 2026-08-20):
- finance (sheet Table 1): B一级子类 C二级子类 D二级定义 E三级子类 F三级定义
  G四级子类 H内容 I安全级别 J备注 K部门意见.
- shougang (sheet 数据分类分级): B一级分类 C一级定义 D二级分类 E二级定义
  F三级分类(G(三级定义) H四级分类 I四级定义 J数据资源说明 K分级 L数据来源.
  A "——"/empty 四级 cell means the leaf lives at 三级 (code carried by G).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping

from agent.standards.contracts import clean

FINANCE_SHEET = "Table 1"
SHOUGANG_SHEET = "数据分类分级"

# finance columns (1-based): letter -> semantic key
FINANCE_COLS = {
    "L1": "B", "L2": "C", "L2_DEF": "D", "L3": "E", "L3_DEF": "F",
    "LEAF": "G", "DESC": "H", "LEVEL": "I", "REMARK": "J", "OPINION": "K",
}
SHOUGANG_COLS = {
    "L1": "B", "L1_DEF": "C", "L2": "D", "L2_DEF": "E", "L3": "F",
    "L3_DEF": "G", "LEAF": "H", "LEAF_DEF": "I", "CONTENT": "J",
    "LEVEL": "K", "RESOURCE": "L",
}


@dataclass(frozen=True)
class CellInfo:
    """One cell resolved through the merged-grid: value + scope provenance."""

    value: Any = None
    anchor_cell: str = ""
    merged_range: str | None = None
    start_row: int | None = None
    end_row: int | None = None
    inherited: bool = False

    def to_mapping(self) -> dict[str, Any]:
        return {
            "value": self.value,
            "anchor_cell": self.anchor_cell,
            "merged_range": self.merged_range,
            "start_row": self.start_row,
            "end_row": self.end_row,
            "inherited": self.inherited,
        }


class MergedCellResolver:
    """Resolve any cell of a worksheet with merged-range awareness.

    Non-anchor cells inside a merged range return the ANCHOR's value plus the
    merged scope (anchor_cell / merged_range / start..end / inherited=True).
    Plain cells return their own value with inherited=False and no range.
    """

    def __init__(self, workbook, sheet: str):
        if sheet not in workbook.sheetnames:
            raise ValueError(f"sheet {sheet!r} not found")
        self._sheet = workbook[sheet]
        self._ranges = list(self._sheet.merged_cells.ranges)

    def close(self) -> None:
        if self._sheet is not None:
            pass  # workbook managed by caller

    def cell(self, row: int, col) -> CellInfo:
        """``col`` is a 1-based integer or a column letter (e.g. 'J')."""
        from openpyxl.utils import column_index_from_string, get_column_letter

        if isinstance(col, str):
            col = column_index_from_string(col)
        cell = self._sheet.cell(row=row, column=col)
        for rng in self._ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                anchor_value = self._sheet.cell(rng.min_row, rng.min_col).value
                return CellInfo(
                    value=anchor_value,
                    anchor_cell=f"{get_column_letter(rng.min_col)}{rng.min_row}",
                    merged_range=str(rng),
                    start_row=rng.min_row,
                    end_row=rng.max_row,
                    inherited=not (row == rng.min_row and col == rng.min_col),
                )
        return CellInfo(
            value=cell.value,
            anchor_cell=f"{get_column_letter(col)}{row}",
            inherited=False,
        )


@dataclass(frozen=True)
class RawEntry:
    """One leaf row of a raw standard, with its traceable origin and the
    hierarchy/annotation fields it INHERITS from merged groups."""

    level_1: str = ""
    level_2: str = ""
    level_3: str = ""
    leaf: str = ""
    description: str = ""
    content: str = ""
    raw_level: str = ""
    resource: str = ""
    level_1_definition: str = ""
    level_2_definition: str = ""
    level_3_definition: str = ""
    remark: str = ""
    department_opinion: str = ""
    sheet: str = ""
    row: int | None = None
    provenance: Mapping[str, Mapping[str, Any]] = field(default_factory=dict)

    def to_mapping(self) -> dict[str, Any]:
        return {
            "level_1": self.level_1,
            "level_2": self.level_2,
            "level_3": self.level_3,
            "leaf": self.leaf,
            "description": self.description,
            "content": self.content,
            "raw_level": self.raw_level,
            "resource": self.resource,
            "level_1_definition": self.level_1_definition,
            "level_2_definition": self.level_2_definition,
            "level_3_definition": self.level_3_definition,
            "remark": self.remark,
            "department_opinion": self.department_opinion,
            "sheet": self.sheet,
            "row": self.row,
            "provenance": dict(self.provenance),
        }


@dataclass(frozen=True)
class ReaderResult:
    entries: tuple[RawEntry, ...] = ()
    issues: tuple[str, ...] = ()


def _open_xlsx(path: Path):
    """Load a workbook with cached values (data_only) for merged-range reads."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("reading raw standard workbooks requires openpyxl") from exc
    return openpyxl.load_workbook(path, data_only=True)


def _prov(cell_info: CellInfo) -> Mapping[str, Any]:
    return {
        "source_cell": cell_info.anchor_cell,
        "merged_range": cell_info.merged_range,
        "start_row": cell_info.start_row,
        "end_row": cell_info.end_row,
        "inherited": cell_info.inherited,
    }


def read_finance_standard_guide(path: str | Path) -> ReaderResult:
    """Extract leaf rows of the finance grading-standard guide (merge-aware)."""
    workbook = _open_xlsx(Path(path))
    resolver = MergedCellResolver(workbook, FINANCE_SHEET)
    entries: list[RawEntry] = []
    issues: list[str] = []
    try:
        for row in range(3, (resolver._sheet.max_row or 2) + 1):
            leaf = clean(resolver.cell(row, FINANCE_COLS["LEAF"]).value)
            if not leaf:
                continue  # trailing rows / headers; not a leaf
            l1i = resolver.cell(row, FINANCE_COLS["L1"])
            l2i = resolver.cell(row, FINANCE_COLS["L2"])
            l2di = resolver.cell(row, FINANCE_COLS["L2_DEF"])
            l3i = resolver.cell(row, FINANCE_COLS["L3"])
            l3di = resolver.cell(row, FINANCE_COLS["L3_DEF"])
            li = resolver.cell(row, FINANCE_COLS["DESC"])
            lvi = resolver.cell(row, FINANCE_COLS["LEVEL"])
            ri = resolver.cell(row, FINANCE_COLS["REMARK"])
            oi = resolver.cell(row, FINANCE_COLS["OPINION"])
            entries.append(
                RawEntry(
                    level_1=clean(l1i.value),
                    level_2=clean(l2i.value),
                    level_3=clean(l3i.value),
                    leaf=leaf,
                    description=clean(li.value),
                    raw_level=str(lvi.value).strip() if lvi.value is not None else "",
                    level_2_definition=clean(l2di.value),
                    level_3_definition=clean(l3di.value),
                    remark=clean(ri.value),
                    department_opinion=clean(oi.value),
                    sheet=FINANCE_SHEET,
                    row=row,
                    provenance={
                        "level_2_definition": _prov(l2di),
                        "level_3_definition": _prov(l3di),
                        "remark": _prov(ri),
                        "department_opinion": _prov(oi),
                    },
                )
            )
    finally:
        workbook.close()
    return ReaderResult(tuple(entries), tuple(issues))


def read_guanji_catalog(path: str | Path) -> ReaderResult:
    """Extract leaf rows of the shougang (关基) grading catalog (merge-aware).

    A \"——\"/empty 四级 cell means the leaf lives at 三级 (the code/name is
    carried by the 三级 column, which is itself a merged group cell).
    """
    workbook = _open_xlsx(Path(path))
    resolver = MergedCellResolver(workbook, SHOUGANG_SHEET)
    entries: list[RawEntry] = []
    issues: list[str] = []
    try:
        for row in range(3, (resolver._sheet.max_row or 2) + 1):
            l1i = resolver.cell(row, SHOUGANG_COLS["L1"])
            l2i = resolver.cell(row, SHOUGANG_COLS["L2"])
            l3i = resolver.cell(row, SHOUGANG_COLS["L3"])
            leaf_h = resolver.cell(row, SHOUGANG_COLS["LEAF"])
            leaf_h_def = resolver.cell(row, SHOUGANG_COLS["LEAF_DEF"])
            if (leaf_h.value is None or clean(leaf_h.value) in ("", "——")):
                # leaf sits at 三级: name+code and its definition come from G
                if l3i.value is None or not clean(l3i.value):
                    issues.append(f"shougang row {row}: no real leaf level")
                    continue
                leaf = clean(l3i.value)
                description = clean(
                    resolver.cell(row, SHOUGANG_COLS["L3_DEF"]).value
                )
            else:
                leaf = clean(leaf_h.value)
                description = clean(leaf_h_def.value)
            content = clean(resolver.cell(row, SHOUGANG_COLS["CONTENT"]).value)
            _level_cell = resolver.cell(row, SHOUGANG_COLS["LEVEL"]).value
            raw_level = str(_level_cell).strip() if _level_cell is not None else ""
            resource = clean(resolver.cell(row, SHOUGANG_COLS["RESOURCE"]).value)
            l1di = resolver.cell(row, SHOUGANG_COLS["L1_DEF"])
            l2di = resolver.cell(row, SHOUGANG_COLS["L2_DEF"])
            l3di = resolver.cell(row, SHOUGANG_COLS["L3_DEF"])
            lresi = resolver.cell(row, SHOUGANG_COLS["RESOURCE"])
            entries.append(
                RawEntry(
                    level_1=clean(l1i.value),
                    level_2=clean(l2i.value),
                    level_3=clean(l3i.value),
                    leaf=leaf,
                    description=description,
                    content=content,
                    raw_level=raw_level,
                    resource=resource,
                    level_1_definition=clean(l1di.value),
                    level_2_definition=clean(l2di.value),
                    level_3_definition=clean(l3di.value),
                    sheet=SHOUGANG_SHEET,
                    row=row,
                    provenance={
                        "level_1_definition": _prov(l1di),
                        "level_2_definition": _prov(l2di),
                        "level_3_definition": _prov(l3di),
                        "resource": _prov(lresi),
                    },
                )
            )
    finally:
        workbook.close()
    return ReaderResult(tuple(entries), tuple(issues))


__all__ = [
    "CellInfo",
    "MergedCellResolver",
    "RawEntry",
    "ReaderResult",
    "read_finance_standard_guide",
    "read_guanji_catalog",
]
